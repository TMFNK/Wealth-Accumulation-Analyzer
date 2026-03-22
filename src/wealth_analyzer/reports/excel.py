"""Excel report generation using openpyxl."""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from wealth_analyzer.config import AppConfig

logger = logging.getLogger(__name__)

_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FONT_GREEN = Font(name="Calibri", size=11, color="276221")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FONT_RED = Font(name="Calibri", size=11, color="9C0006")
_FILL_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _write_header(ws, columns: list[str]) -> None:
    """Write a formatted header row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")


def _auto_fit_columns(ws) -> None:
    """Auto-fit column widths to content, capped at 40."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_data_rows(ws, df: pd.DataFrame, start_row: int = 2) -> None:
    """Write DataFrame rows starting at start_row with alternating shading."""
    for row_idx, (_, row) in enumerate(df.iterrows(), start_row):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Handle NaN/None
            if val is None or (isinstance(val, float) and np.isnan(val)):
                cell.value = None
            else:
                cell.value = val
            if (row_idx - start_row) % 2 == 1:
                cell.fill = _FILL_ALT


def _apply_conditional_format(
    ws, col_name: str, columns: list[str], start_row: int, end_row: int
) -> None:
    """Apply green/red conditional formatting to a column."""
    if col_name not in columns:
        return
    col_idx = columns.index(col_name) + 1
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        val = cell.value
        if isinstance(val, (int, float)) and not np.isnan(val):
            if val > 0:
                cell.fill = _FILL_GREEN
                cell.font = _FONT_GREEN
            elif val < 0:
                cell.fill = _FILL_RED
                cell.font = _FONT_RED


def _format_value_columns(ws, columns: list[str], start_row: int, end_row: int) -> None:
    """Apply number formatting to specific column patterns."""
    for col_idx, col_name in enumerate(columns, 1):
        fmt = None
        if col_name.endswith("_pct") or col_name.endswith("_pp"):
            fmt = "0.00%"
        elif "value" in col_name.lower() or "contributions" in col_name.lower():
            fmt = "#,##0.00"
        elif col_name == "years_analyzed":
            fmt = "0.0"
        if fmt:
            for row in range(start_row, end_row + 1):
                ws.cell(row=row, column=col_idx).number_format = fmt


def write_excel(
    lump_sum_results: pd.DataFrame,
    lump_sum_growth: pd.DataFrame,
    dca_results: pd.DataFrame,
    dca_growth: pd.DataFrame,
    dca_cost_basis: pd.DataFrame,
    cfg: AppConfig,
) -> Path:
    """Write a multi-sheet Excel workbook.

    Parameters
    ----------
    lump_sum_results, lump_sum_growth:
        Outputs from ``run_lump_sum()``.
    dca_results, dca_growth, dca_cost_basis:
        Outputs from ``run_dca()``.
    cfg:
        Application configuration.

    Returns
    -------
    Path
        The path to the written workbook.
    """
    output_dir = Path(cfg.output.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = cfg.output.excel_filename.replace("{date}", date.today().isoformat())
    path = output_dir / filename

    wb = Workbook()

    # --- Sheet 1: Lump-Sum Summary ---
    ws_ls = wb.active
    ws_ls.title = "Lump-Sum Summary"
    if not lump_sum_results.empty:
        cols = list(lump_sum_results.columns)
        _write_header(ws_ls, cols)
        _write_data_rows(ws_ls, lump_sum_results)
        end_row = len(lump_sum_results) + 1
        _apply_conditional_format(ws_ls, "outperformance_cagr_pp", cols, 2, end_row)
        _format_value_columns(ws_ls, cols, 2, end_row)
    ws_ls.freeze_panes = "A2"
    _auto_fit_columns(ws_ls)

    # --- Sheet 2: DCA Summary ---
    ws_dca = wb.create_sheet("DCA Summary")
    if not dca_results.empty:
        cols = list(dca_results.columns)
        _write_header(ws_dca, cols)
        _write_data_rows(ws_dca, dca_results)
        end_row = len(dca_results) + 1
        _apply_conditional_format(ws_dca, "xirr_pct", cols, 2, end_row)
        _format_value_columns(ws_dca, cols, 2, end_row)
    ws_dca.freeze_panes = "A2"
    _auto_fit_columns(ws_dca)

    # --- Per-ticker sheets ---
    all_tickers = set()
    if not lump_sum_growth.empty:
        all_tickers.update(lump_sum_growth.columns)
    if not dca_growth.empty:
        all_tickers.update(dca_growth.columns)

    for ticker in sorted(all_tickers):
        ws_t = wb.create_sheet(ticker)
        ws_t.cell(row=1, column=1, value="Date").font = Font(bold=True)
        ws_t.cell(row=1, column=2, value="Lump-Sum Value").font = Font(bold=True)
        ws_t.cell(row=1, column=3, value="DCA Value").font = Font(bold=True)
        ws_t.cell(row=1, column=4, value="DCA Cost Basis").font = Font(bold=True)
        for c in range(1, 5):
            ws_t.cell(row=1, column=c).fill = _FILL_HEADER
            ws_t.cell(row=1, column=c).font = _FONT_HEADER

        # Collect all dates
        ls_series = (
            lump_sum_growth[ticker]
            if ticker in lump_sum_growth.columns
            else pd.Series(dtype=float)
        )
        dca_series = (
            dca_growth[ticker]
            if ticker in dca_growth.columns
            else pd.Series(dtype=float)
        )
        cb_series = (
            dca_cost_basis[ticker]
            if ticker in dca_cost_basis.columns
            else pd.Series(dtype=float)
        )

        all_idx = (
            ls_series.index.union(dca_series.index).union(cb_series.index).sort_values()
        )

        for row_idx, dt in enumerate(all_idx, 2):
            ws_t.cell(row=row_idx, column=1, value=dt).number_format = "YYYY-MM-DD"
            lv = ls_series.get(dt)
            dv = dca_series.get(dt)
            cv = cb_series.get(dt)
            ws_t.cell(
                row=row_idx, column=2, value=float(lv) if pd.notna(lv) else None
            ).number_format = "#,##0.00"
            ws_t.cell(
                row=row_idx, column=3, value=float(dv) if pd.notna(dv) else None
            ).number_format = "#,##0.00"
            ws_t.cell(
                row=row_idx, column=4, value=float(cv) if pd.notna(cv) else None
            ).number_format = "#,##0.00"

        last_row = max(len(all_idx) + 1, 2)

        # Embed chart if we have data
        if len(all_idx) > 0:
            chart = LineChart()
            chart.title = f"{ticker} \u2014 Portfolio Growth"
            chart.width = 20  # cm
            chart.height = 12  # cm
            chart.y_axis.title = "Value (USD)"
            chart.x_axis.title = "Date"

            cats = Reference(ws_t, min_col=1, min_row=2, max_row=last_row)
            for col_idx in [2, 3, 4]:
                data = Reference(ws_t, min_col=col_idx, min_row=1, max_row=last_row)
                chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.smooth = True

            # Style series
            if len(chart.series) >= 1:
                chart.series[0].graphicalProperties.line.solidFill = "1F77B4"
            if len(chart.series) >= 2:
                chart.series[1].graphicalProperties.line.solidFill = "FF7F0E"
            if len(chart.series) >= 3:
                chart.series[2].graphicalProperties.line.dashStyle = "dash"
                chart.series[2].graphicalProperties.line.solidFill = "2CA02C"

            ws_t.add_chart(chart, "F2")

        _auto_fit_columns(ws_t)

    # --- Config sheet ---
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.cell(row=1, column=1, value="Parameter").font = Font(bold=True)
    ws_cfg.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws_cfg.cell(row=1, column=1).fill = _FILL_HEADER
    ws_cfg.cell(row=1, column=1).font = _FONT_HEADER
    ws_cfg.cell(row=1, column=2).fill = _FILL_HEADER
    ws_cfg.cell(row=1, column=2).font = _FONT_HEADER

    params = [
        ("Run Date", date.today().isoformat()),
        ("Start Date", str(cfg.general.start_date)),
        ("End Date", str(cfg.general.end_date)),
        ("Lump Sum Amount", cfg.investment.lump_sum_amount),
        ("DCA Monthly Amount", cfg.investment.dca_monthly_amount),
        ("Risk-Free Rate", cfg.investment.risk_free_rate),
        ("Stocks", ", ".join(cfg.tickers.stocks)),
        ("ETFs", ", ".join(cfg.tickers.etfs)),
        ("Cache TTL Days", cfg.general.cache_ttl_days),
        ("Output Dir", cfg.output.output_dir),
    ]
    for row_idx, (param, val) in enumerate(params, 2):
        ws_cfg.cell(row=row_idx, column=1, value=param)
        ws_cfg.cell(row=row_idx, column=2, value=val)
    _auto_fit_columns(ws_cfg)

    # --- README sheet ---
    ws_readme = wb.create_sheet("README")
    ws_readme.merge_cells("A1:D1")
    cell = ws_readme.cell(
        row=1,
        column=1,
        value=f"Generated by Wealth Accumulation Analyzer \u2014 {date.today().isoformat()}",
    )
    cell.font = Font(bold=True, size=14)
    ws_readme.cell(
        row=3,
        column=1,
        value="This workbook contains lump-sum and DCA analysis results. "
        "Sheets per ticker show daily portfolio values with embedded charts. "
        "See the Config sheet for run parameters.",
    )

    wb.save(str(path))
    logger.info("Excel report written to %s", path)
    return path
