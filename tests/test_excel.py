"""Tests for Excel report generation."""

from __future__ import annotations

from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from wealth_analyzer.config import (
    AppConfig,
    GeneralConfig,
    InvestmentConfig,
    OutputConfig,
    TickersConfig,
)
from wealth_analyzer.reports.excel import write_excel


def _make_config(tmp_path: Path) -> AppConfig:
    return AppConfig(
        general=GeneralConfig(
            start_date=date(2019, 1, 2), end_date=date(2023, 12, 29), cache_ttl_days=1
        ),
        investment=InvestmentConfig(
            lump_sum_amount=10_000, dca_monthly_amount=500, risk_free_rate=0.045
        ),
        tickers=TickersConfig(stocks=["AAPL"], etfs=["SPY"]),
        output=OutputConfig(output_dir=str(tmp_path)),
    )


def _make_data() -> dict:
    dates = pd.bdate_range("2020-01-02", periods=50)
    ls_results = pd.DataFrame(
        [
            {
                "ticker": "AAPL",
                "benchmark": "SPY",
                "cagr": 0.25,
                "sharpe": 1.1,
                "outperformance_cagr_pp": 12.5,
            }
        ]
    )
    ls_growth = pd.DataFrame(
        {
            "AAPL": np.linspace(10_000, 15_000, 50),
            "SPY": np.linspace(10_000, 12_000, 50),
        },
        index=dates,
    )
    dca_results = pd.DataFrame(
        [
            {
                "ticker": "AAPL",
                "total_contributions": 30_000,
                "final_value": 45_000,
                "xirr_pct": 0.12,
                "cagr": 0.20,
            }
        ]
    )
    dca_growth = pd.DataFrame({"AAPL": np.linspace(0, 45_000, 50)}, index=dates)
    dca_cost = pd.DataFrame({"AAPL": np.linspace(0, 30_000, 50)}, index=dates)
    return dict(
        ls_results=ls_results,
        ls_growth=ls_growth,
        dca_results=dca_results,
        dca_growth=dca_growth,
        dca_cost=dca_cost,
    )


class TestWriteExcel:
    def test_creates_file(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        assert path.is_file()

    def test_file_non_empty(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        assert path.stat().st_size > 10_000

    def test_expected_sheet_names(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        wb = load_workbook(path)
        assert "Lump-Sum Summary" in wb.sheetnames
        assert "DCA Summary" in wb.sheetnames
        assert "Config" in wb.sheetnames
        assert "README" in wb.sheetnames

    def test_lump_sum_row_count(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        wb = load_workbook(path)
        ws = wb["Lump-Sum Summary"]
        # 1 header + 1 data row
        assert ws.max_row == 2

    def test_dca_row_count(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        wb = load_workbook(path)
        ws = wb["DCA Summary"]
        assert ws.max_row == 2

    def test_per_ticker_sheet_exists(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        wb = load_workbook(path)
        assert "AAPL" in wb.sheetnames

    def test_config_sheet_rows(self, tmp_path: Path) -> None:
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            _make_config(tmp_path),
        )
        wb = load_workbook(path)
        ws = wb["Config"]
        # 1 header + 10 param rows
        assert ws.max_row >= 10

    def test_creates_output_dir(self, tmp_path: Path) -> None:
        deep = tmp_path / "deep" / "nested"
        cfg = _make_config(deep)
        d = _make_data()
        path = write_excel(
            d["ls_results"],
            d["ls_growth"],
            d["dca_results"],
            d["dca_growth"],
            d["dca_cost"],
            cfg,
        )
        assert path.is_file()
        assert deep.is_dir()
